import openpyxl
from pathlib import Path
from typing import Dict, Any, List

from ..utils.logging_config import get_logger

logger = get_logger(__name__)


class ExcelParser:
    """
    Parser для Excel документов
    
    Extracts:
    - All sheets data
    - Formulas
    - Cell formatting
    - Metadata
    """
    
    def parse(self, excel_path: Path) -> Dict[str, Any]:
        """
        Парсинг Excel файла
        
        Args:
            excel_path: Путь к Excel
        
        Returns:
            Dict со всеми листами и данными
        """
        logger.info(f"Parsing Excel: {excel_path}")
        
        result = {
            'sheets': {},
            'formulas': [],
            'metadata': {},
            'file_size': excel_path.stat().st_size
        }
        
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=False)
            
            result['metadata'] = {
                'creator': wb.properties.creator,
                'created': str(wb.properties.created),
                'modified': str(wb.properties.modified),
                'num_sheets': len(wb.sheetnames)
            }
            
            # Парсинг каждого листа
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Извлечение данных
                sheet_data = []
                for row in sheet.iter_rows(values_only=True):
                    sheet_data.append(list(row))
                
                result['sheets'][sheet_name] = sheet_data
                
                # Извлечение формул
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            if cell.value.startswith('='):
                                result['formulas'].append({
                                    'sheet': sheet_name,
                                    'cell': cell.coordinate,
                                    'formula': cell.value
                                })
            
            logger.info(
                f"✅ Excel parsed: {len(result['sheets'])} sheets, "
                f"{len(result['formulas'])} formulas"
            )
            
        except Exception as e:
            logger.error(f"Excel parsing error: {e}", exc_info=True)
            raise
        
        return result
